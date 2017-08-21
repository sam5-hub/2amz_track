# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html

from scrapy.utils.project import get_project_settings
from scrapy.pipelines.images import ImagesPipeline
import os
import platform
from datetime import datetime
from openpyxl import Workbook

settings = get_project_settings()
all_resource_path = settings.attributes.get('RESOURCE_STORE').value
resource_path = settings.attributes.get('IMAGES_STORE').value
today_path = os.path.join(all_resource_path, datetime.now().strftime('%y-%m-%d'))


class AmzCrawlXLSXPipeline(object):
    wb = Workbook()

    def process_item(self, item, spider):  # 工序具体内容
        self.insert_item_to_xlsx(item)
        return item

    def insert_item_to_xlsx(self, item):
        category = item['category']
        line = [category, item['sub_category'], item['title'], item['rank']]  # 把数据中每一项整理出来
        belong_dir = os.path.join(today_path, item['belong'])
        if not os.path.exists(belong_dir):
            os.makedirs(belong_dir)

        xlsx_path = os.path.join(today_path, datetime.now().strftime('%y-%m-%d') + '.xlsx')

        ws = self.wb.active

        if category in self.wb.sheetnames:
            ws = self.wb.get_sheet_by_name(category)
            ws.append(line)

        if category not in self.wb.sheetnames:
            if ws.title == 'Sheet':
                ws.title = category
            else:
                ws = self.wb.create_sheet(title=category)
            ws.append(['分类', '子分类', '标题', '排名'])  # 设置表头
            ws.append(line)  # 将数据以行的形式添加到xlsx中

            self.wb.save(xlsx_path)  # 保存xlsx文件


class AmzCrawlPipeline(object):
    def process_item(self, item, spider):
        self.process_file(item, spider)
        return item

    # 文件夹
    def process_file(self, item, spider):
        today_dir = os.path.join(resource_path, datetime.now().strftime('%y-%m-%d'))
        belong_dir = os.path.join(today_dir, item['belong'])
        sub_category_dir = os.path.join(belong_dir, item['sub_category'])

        if not os.path.exists(sub_category_dir):
            os.makedirs(sub_category_dir, exist_ok=True)


class AmazonOrdersImagePipeline(ImagesPipeline):
    def item_completed(self, results, item, info):
        if "front_image_url" in item:
            for ok, value in results:
                if platform.system() == 'Windows':  # full 图片位置
                    image_file_path = value["path"].replace('/', '\\')
                else:
                    image_file_path = value["path"]

                old_path = os.path.join(resource_path, image_file_path)

                if os.path.exists(old_path):
                    self.move_old_new(image_file_path, item)

        return item

    def move_old_new(self, image_file_path, item):

        today_dir = os.path.join(resource_path, datetime.now().strftime('%y-%m-%d'))
        belong_dir = os.path.join(today_dir, item['belong'])
        sub_category_dir = os.path.join(belong_dir, item['sub_category'])

        if not os.path.exists(sub_category_dir):
            os.makedirs(sub_category_dir, exist_ok=True)
        new_path = os.path.join(sub_category_dir, item['rank'] + '.' + item['des'] + '.jpg')
        old_path = os.path.join(resource_path, image_file_path)

        return new_path, old_path
